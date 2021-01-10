from openpyxl import Workbook
from openpyxl import load_workbook

# # 读取整个excel
# wb = load_workbook(filename='./sample.xlsx')
# # 读取excel中worksheet
# ws = wb.active
# print(ws['A1'].value)
# # 读取每一行
# for row in ws.values:
#     print('row：', row)


# 实例化workbook相当于打开excel文件
wb = Workbook()
# 打开默认的worksheet
ws = wb.active

ws['A1'] = 'helloworld'
ws['B3'] = 'xiaoming'
ws.append([1, 2, 3])
# 保存excel文件
wb.save('./datadriven.xlsx')

